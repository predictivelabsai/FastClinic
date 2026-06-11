"""Spreadsheet export helpers — real .xlsx downloads with a FastClinic styled header.

Mirrors the pattern used in sister repos: build an openpyxl Workbook in memory
and return it as a Starlette Response with the xlsx MIME type.
"""
from __future__ import annotations

import io

from starlette.responses import Response

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# FastClinic primary-blue header.
_HEADER_BG = "1E6FB8"
_HEADER_FG = "FFFFFF"


def build_xlsx(headers: list[str], rows: list[list], sheet_name: str = "Export") -> bytes:
    """Build a single-sheet .xlsx with a styled header row and return the bytes."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Export")[:31]  # Excel sheet-name limit

    ws.append([str(h) for h in headers])
    fill = PatternFill(start_color=_HEADER_BG, end_color=_HEADER_BG, fill_type="solid")
    font = Font(color=_HEADER_FG, bold=True, size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    widths = [max(10, len(str(h)) + 2) for h in headers]
    for r in rows:
        ws.append([_cell(v) for v in r])
        for i, v in enumerate(r):
            if i < len(widths):
                widths[i] = min(60, max(widths[i], len(str("" if v is None else v)) + 2))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(v):
    if isinstance(v, (list, tuple)):
        return "; ".join(str(x) for x in v)
    return v


def xlsx_response(filename: str, headers: list[str], rows: list[list],
                  sheet_name: str = "Export") -> Response:
    return Response(
        content=build_xlsx(headers, rows, sheet_name),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
